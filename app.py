import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# 读取售后数据文件
df = pd.read_excel(r'C:\Users\Administrator\Desktop\PY\售后数据处理\售后数据.xlsx')

# 删除"故障部位标签"列中空白单元格的所有行
df.dropna(subset=['故障部位标签'], inplace=True)

# 只保留指定列并重命名
df = df[['物料组', 'SN扫码结果', '服务工单类型', '产品系列', '创建时间', '故障部位标签', '故障部位', '故障码', '故障现象']]
df.rename(columns={'物料组': '产品类型'}, inplace=True)

# 删除"SN扫码结果"列中首字母非R的行
df = df[df['SN扫码结果'].str.startswith('R')]


# 提取年份和周数，确保格式为 25-01 等
df['故障周数'] = df['创建时间'].dt.strftime('%y') + '-' + df['创建时间'].dt.strftime('%U').str.zfill(2)


# 格式化"创建时间"列，只保留年/月，日期格式为"25-01"
df['创建时间'] = df['创建时间'].dt.strftime('%y-%m')

# 在"问题产生时间"列后插入新列"故障数"，并填充为1
df.insert(df.columns.get_loc('故障部位标签') + 1, '故障数', 1)

# 提取"SN扫码结果"列的第8-9个字符，并创建新列"生产批次"
df['生产批次'] = '2' + df['SN扫码结果'].str[6:7] + '-' + df['SN扫码结果'].str[7:9]

# 只保留前缀为24或25的生产批次
df = df[df['生产批次'].str.startswith(('24-', '25-'))]


# 读取销量数据文件
sales_df = pd.read_excel(r'C:\Users\Administrator\Desktop\PY\售后数据处理\销量数据.xlsx')

# 创建新列"累计销量"，并根据"产品系列"匹配销量数据
df['累计销量'] = df['产品系列'].map(sales_df.set_index('产品系列')['销量'])

# 修改"产品系列"列内容，根据销量数据中的A列和C列进行匹配
df['产品系列'] = df['产品系列'].map(sales_df.set_index('产品系列')['产品名称'])

# 将数据按"产品类型"分为两个Sheet表
df_robot = df[df['产品类型'] == '产品_扫地机器人']
df_cleaner = df[df['产品类型'] == '产品_家用洗地机']

# 保存结果到Excel文件
with pd.ExcelWriter(r'C:\Users\Administrator\Desktop\PY\售后数据处理\数据处理.xlsx', engine='openpyxl') as writer:
    df_robot.to_excel(writer, sheet_name='产品_扫地机器人', index=False)
    df_cleaner.to_excel(writer, sheet_name='产品_家用洗地机', index=False)
    
    # 设置单元格格式
    workbook = writer.book
    for sheet_name in writer.sheets:
        worksheet = writer.sheets[sheet_name]
        
        # 设置列宽为20
        for col in worksheet.columns:
            column_letter = get_column_letter(col[0].column)
            worksheet.column_dimensions[column_letter].width = 20
        
        # 设置表头为蓝色字体
        for cell in worksheet[1]:
            cell.font = Font(color='0000FF', bold=True)
        
        # 设置所有单元格文字居中
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # 设置日期格式
                if cell.column_letter in ['故障周数', '创建时间']:
                    cell.number_format = 'YYYY-MM'

print("数据处理完成并保存成功！")